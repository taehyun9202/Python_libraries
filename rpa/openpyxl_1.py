from openpyxl import Workbook
wb = Workbook()

# ws = wb.active 

ws = wb.create_sheet()
ws.title = "RPA Sheet"
ws.sheet_properties.tabColor = '000000'

ws1 = wb.create_sheet('another sheet')
ws2 = wb.create_sheet('new sheet', 2)

new_ws = wb['RPA Sheet']



# sheet copy

new_ws['A1'] = "TEST"
target = wb.copy_worksheet(new_ws)
target.title = "copied sheet"


print(wb.sheetnames)
wb.save('sample.xlsx')
# wb.close()



