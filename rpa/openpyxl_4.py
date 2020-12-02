from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
wb =load_workbook('sample.xlsx')
ws = wb.active


for row in ws.iter_rows(min_row=2):
    #num, eng, math
    if int(row[1].value) > 80:
        print(row[0].value)

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "English":
            cell.value = "Computer"

# ws.insert_rows(8, 5) # insert 5 lines at line 8
# ws.insert_cols(2, 3) # insert 3 rows at row 2

# ws.delete_rows(8, 3) # delete 3 rows from row 8
# ws.delete_cols(2, 2) # delte tow col from col 2

# num, computer, math  
# num, english, computer, math

# ws.move_range("B1:C11", rows=0, cols=1)   # grab data from b1 to c11
# ws['B1'].value = "English"

# bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
# bar_chart = BarChart()
# bar_chart.add_data(bar_value)
# ws.add_chart(bar_chart, "E1")  

line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3)
line_chart = LineChart()
line_chart.add_data(line_value, titles_from_data=True)
line_chart.title = "Grade Card"
line_chart.style = 10
line_chart.y_axis.title = "Grade"
line_chart.x_axis.title = "Number"
ws.add_chart(line_chart, "E1")  

# pie_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
# pie_chart = PieChart()
# pie_chart.add_data(pie_value)
# ws.add_chart(pie_chart, "E1")  

wb.save('sample_modified.xlsx')