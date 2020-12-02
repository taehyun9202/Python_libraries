from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.title ="RPA Sheet"

# ws['A1'] = 1
# ws['A2'] = 2
# ws['A3'] = 3

# ws['b1'] = 4
# ws['b2'] = 5
# ws['b3'] = 6


# print(ws['A1'])
# print(ws['A1'].value)
# print(ws['A10'].value)
# print(ws.cell(row=1, column=1).value)
# print(ws.cell(row=1, column=2).value)

# c = ws.cell(column=3, row=1, value=10) # ws['c1'].value = 10
# print(c.value)



# index = 1
# for x in range(1, 11):
#     for y in range(1, 11):
#         # ws.cell(row=x, column=y, value=randint(0, 100))
#         ws.cell(row=x, column=y, value=index)
#         index += 1

# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()


for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()


wb.save('sample.xlsx')