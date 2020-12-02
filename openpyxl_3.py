from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from random import *

wb = Workbook()

ws = wb.active

ws.append(['Num', 'English', 'Math'])
for i in range(1, 11):
    ws.append([i, randint(0, 101), randint(0, 101)])

# col_eng = ws['B'] 
# print(col_eng)

# for idx, cell in enumerate(col_eng):
#     print(idx, cell.value)

# col_range = ws["B:C"] # column between B and C
# for cols in col_range:
#     for idx, cell in enumerate(cols):
#         if (idx == 0):
#             print(cell.value)
#         else:
#             print(idx, cell.value)


# row_title = ws[1] 
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # row between 2 and 6
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()


# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# print(tuple(ws.rows))

# print(tuple(ws.columns))

# for row in tuple(ws.rows):
#     print(row[0].value)

# for col in tuple(ws.columns):
#     print(col[0].value)

# for row in ws.iter_rows():
#     print(row[2].value)

# for col in ws.iter_cols():
#     print(col[0].value)

# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    # print(row[0].value, row[1].value)
    # print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save('sample.xlsx')