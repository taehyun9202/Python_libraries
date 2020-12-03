import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1, 2, 3)"
ws["A3"] = "=AVERAGE(1, 2, 3)"
a = 5
b = 10
c = 16
ws["A4"] = "=SUM(A2:A3)"

ws["A6"] = "=SUM({}, {}, {})".format(a, b, c)
ws["A7"] = f"=AVERAGE({a}, {b}, {c})"


# print formula
    # for row in ws.values:
    #     for cell in row:
    #         print(cell)


# print data  (need to save excel to see actual data)
# wb = load_workbook("sample_formula.xlsx", data_only=True)
# for row in ws.values:
#     for cell in row:
#         print(cell)

######################## MERgE #######################

ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.unmerge_cells("B2:D2")


img = Image('image.jpg')

ws.add_image(img, "C3")

wb.save("sample2.xlsx")