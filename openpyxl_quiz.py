from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = Workbook()

wb = Workbook()

ws = wb.active


ws.append(("Number", "Attendence", "Quiz 1", "Quiz 2", "Mid Term", "Final", "Project"))

scores = [
    (1, 10, 8, 5, 14, 26, 12),
    (2, 7, 3, 7, 15, 24, 18),
    (3, 9, 5, 5, 8, 12, 12),
    (4, 7, 8, 5, 19, 21, 18),
    (5, 7, 8, 5, 16, 25, 15),
    (6, 3, 5, 9, 8, 17, 8),
    (7, 4, 9, 5, 14, 26, 18),
    (8, 6, 6, 6, 15, 19, 17),
    (9, 10, 10, 9, 19, 30, 19),
    (10, 9, 8, 8, 20, 25, 20),
]

for idx, s in enumerate(scores):
    ws.append(s)


# change all quiz 2 score to 10
for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10

# add total score in H and grade in H
ws["H1"] = "Total"
ws["I1"] = "Grade"

row_range = ws["2:11"]
total = 0

wb.save('Scores.xlsx') 

wb = load_workbook('Scores.xlsx', data_only=True)
ws = wb.active

for idx, row in enumerate(row_range, start=2):
    ws.cell(row=idx, column=8).value = f"=SUM(B{idx}:G{idx})"
    total = ws.cell(row=idx, column=8).value
    print(total)

    if int(ws.cell(row=idx, column=2).value) < 5:
        grade = "F"
    elif total > 90:
        grade = "A"
    elif total >= 80:
        grade = "B"
    elif total >= 70:
        grade = "C"
    else:
        grade = "D"

    ws.cell(row=idx, column=9).value = grade
